#
# This file is part of the BIOM_AID distribution (https://bitbucket.org/kig13/dem/).
# Copyright (c) 2020-2021 Brice Nord, Romuald Kliglich, Alexandre Jaborska, Philomène Mazand.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, version 3.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program. If not, see <http://www.gnu.org/licenses/>.
#
import re

import pandas as pd
from django.core.management.base import BaseCommand
from django.utils.translation import gettext_lazy as _

import openpyxl

from common import config

cmd_re = re.compile(r"((?:if|IF|ii|II)\s?\d\d\d\d\d\d)")
intv_re = re.compile(r"\b(\d\d\d\d\d.)\b")


class Command(BaseCommand):
    help = "Traite une extraction de Prodige (factures) et tente de rapprocher avec les commandes magh2 et/ou les"
    " interventions d'Asset+"

    def add_arguments(self, parser):
        parser.add_argument('filename', help=_("Nom du fichier Prodige (format xlsx)"))

    def handle(self, *args, **options):
        self.stdout.write(self.style.NOTICE(_("Analyse fichier Prodige : '{:s}' ...").format(options['filename'])))
        try:
            wb = openpyxl.load_workbook(options['filename'])
            ws = wb[wb.sheetnames[0]]  # TODO: A mettre dans la config et/ou en option (feuille à utiliser)
            titles = {}
            commande_col = None
            fact_commandes = {}
            for i_col in range(1, ws.max_column):
                titles[i_col] = ws.cell(1, i_col).value
                if ws.cell(1, i_col).value == config.get('facturation')['prodige']['colonne_commande']:
                    commande_col = i_col
            for i_row in range(2, ws.max_row + 1):
                for i_col in range(1, ws.max_column):
                    if i_col == commande_col:
                        for cmd in (cmd.upper().replace(' ', '') for cmd in cmd_re.findall(ws.cell(i_row, i_col).value)):
                            # print(i_row - 1, cmd)
                            fact_commandes[cmd] = {
                                ws.cell(1, i_col).value: ws.cell(i_row, i_col).value for i_col in range(1, ws.max_column)
                            }
                            ...
        except FileNotFoundError:
            self.stdout.write(self.style.Error("Fichier '{:s}' non trouvé.").format(options['filename']))

        self.stdout.write(self.style.NOTICE(_("  {:d} factures avec commande trouvées.").format(len(fact_commandes))))

        commandes = pd.read_csv(
            config.get('facturation')['commandes_c6']['path'] + '/' + config.get('facturation')['commandes_c6']['filename']
        )
        print(commandes.columns)
        commandes['no_cmd'] = commandes['Gest. (ec)'] + commandes['No Cde (ec)'].apply(lambda v: '{:06d}'.format(v))
        commandes['bn'] = commandes['Bloc Note 1 (ec)'].fillna('') + commandes['Bloc Note 2 (ec)'].fillna('')
        for c in commandes.iterrows():
            if c[1]['no_cmd'] in fact_commandes.keys():
                bn = c[1]['bn']
                print(bn, intv_re.findall(bn))

        self.stdout.write(self.style.SUCCESS("Terminé."))
